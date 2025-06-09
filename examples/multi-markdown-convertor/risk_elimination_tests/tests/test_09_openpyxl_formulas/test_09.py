#!/usr/bin/env python3
"""
Test 9: `openpyxl` formula string vs. cached value parity
XLSX with volatile formulas
"""

import json
import sys
import tempfile
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime


def create_test_xlsx() -> Path:
    """Create a test XLSX file with volatile formulas."""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Formula Test"

        # Add headers
        ws["A1"] = "Formula Type"
        ws["B1"] = "Formula"
        ws["C1"] = "Expected Result"

        # Add volatile formulas
        formulas = [
            ("NOW()", "=NOW()", "Current date/time"),
            ("TODAY()", "=TODAY()", "Current date"),
            ("RAND()", "=RAND()", "Random number"),
            ("RANDBETWEEN()", "=RANDBETWEEN(1,100)", "Random 1-100"),
            ("Simple Math", "=2+3", "5"),
            ("Cell Reference", "=A1", "Formula Type"),
        ]

        for i, (desc, formula, expected) in enumerate(formulas, start=2):
            ws[f"A{i}"] = desc
            ws[f"B{i}"] = formula
            ws[f"C{i}"] = expected

        # Save to temp file
        temp_file = Path(tempfile.mktemp(suffix=".xlsx"))
        wb.save(str(temp_file))
        return temp_file

    except ImportError:
        # Fallback: use existing XLSX file if available
        test_inputs_dir = Path(__file__).parent.parent.parent / "test-inputs"
        xlsx_files = list(test_inputs_dir.glob("*.xlsx"))
        if xlsx_files:
            return xlsx_files[0]
        else:
            raise RuntimeError(
                "openpyxl not available and no test XLSX files found"
            )


def analyze_formula_values(xlsx_path: Path) -> Dict[str, Any]:
    """Analyze formula strings vs cached values in XLSX."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(xlsx_path), data_only=False)
        wb_data_only = load_workbook(str(xlsx_path), data_only=True)

        analysis = {
            "worksheets_found": len(wb.worksheets),
            "formulas_found": 0,
            "formula_details": [],
            "mismatches": 0,
            "mismatch_rate": 0.0,
        }

        for ws_name in wb.sheetnames:
            ws_formula = wb[ws_name]
            ws_data = wb_data_only[ws_name]

            for row in ws_formula.iter_rows():
                for cell in row:
                    if (
                        cell.value
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        analysis["formulas_found"] += 1

                        # Get corresponding data cell
                        data_cell = ws_data[cell.coordinate]

                        formula_info = {
                            "cell": cell.coordinate,
                            "formula": cell.value,
                            "cached_value": data_cell.value,
                            "formula_value": cell.value,  # This is the formula string
                            "is_volatile": is_volatile_formula(cell.value),
                            "has_mismatch": False,
                        }

                        # For non-volatile formulas, we can check if values make sense
                        if not formula_info["is_volatile"]:
                            # Simple check for basic math formulas
                            if cell.value == "=2+3" and data_cell.value != 5:
                                formula_info["has_mismatch"] = True
                                analysis["mismatches"] += 1

                        analysis["formula_details"].append(formula_info)

        if analysis["formulas_found"] > 0:
            analysis["mismatch_rate"] = (
                analysis["mismatches"] / analysis["formulas_found"]
            )

        return analysis

    except ImportError:
        return {
            "error": "openpyxl not available",
            "worksheets_found": 0,
            "formulas_found": 0,
            "formula_details": [],
            "mismatches": 0,
            "mismatch_rate": 0.0,
        }
    except Exception as e:
        return {
            "error": str(e),
            "worksheets_found": 0,
            "formulas_found": 0,
            "formula_details": [],
            "mismatches": 0,
            "mismatch_rate": 0.0,
        }


def is_volatile_formula(formula: str) -> bool:
    """Check if a formula contains volatile functions."""
    volatile_functions = [
        "NOW()",
        "TODAY()",
        "RAND()",
        "RANDBETWEEN()",
        "INDIRECT()",
        "OFFSET()",
    ]
    formula_upper = formula.upper()
    return any(func in formula_upper for func in volatile_functions)


def run_test() -> Dict[str, Any]:
    """
    Run test 9: Check openpyxl formula string vs cached value parity.

    Returns:
        Dict with test results
    """
    results: Dict[str, Any] = {
        "test_id": "09",
        "test_name": "`openpyxl` formula string vs. cached value parity",
        "test_case": "XLSX with volatile formulas",
        "success": False,
        "error": None,
        "details": {},
    }

    try:
        print(f"Test 9: `openpyxl` formula string vs. cached value parity")
        print(f"Test case: XLSX with volatile formulas")

        # Create or find test XLSX
        print("Creating test XLSX with volatile formulas...")
        xlsx_path = create_test_xlsx()
        results["details"]["xlsx_path"] = str(xlsx_path)

        # Analyze formulas
        print("Analyzing formula strings vs cached values...")
        analysis = analyze_formula_values(xlsx_path)
        results["details"]["formula_analysis"] = analysis

        # Determine success
        if "error" in analysis:
            results["error"] = analysis["error"]
            results["success"] = False
            results["details"]["result"] = f"FAIL: {analysis['error']}"
            print(f"❌ FAIL: {analysis['error']}")
        else:
            if analysis["formulas_found"] > 0:
                mismatch_rate = analysis["mismatch_rate"]
                if mismatch_rate < 0.1:  # Less than 10% mismatch rate
                    results["success"] = True
                    results["details"]["result"] = (
                        f"PASS: Low mismatch rate {mismatch_rate:.1%} ({analysis['mismatches']}/{analysis['formulas_found']})"
                    )
                    print(
                        f"✅ PASS: Low mismatch rate {mismatch_rate:.1%} ({analysis['mismatches']}/{analysis['formulas_found']})"
                    )
                else:
                    results["success"] = False
                    results["details"]["result"] = (
                        f"FAIL: High mismatch rate {mismatch_rate:.1%} ({analysis['mismatches']}/{analysis['formulas_found']})"
                    )
                    print(
                        f"❌ FAIL: High mismatch rate {mismatch_rate:.1%} ({analysis['mismatches']}/{analysis['formulas_found']})"
                    )
            else:
                results["success"] = False
                results["details"]["result"] = "FAIL: No formulas found"
                print("❌ FAIL: No formulas found")

        # Clean up temp file
        if xlsx_path.name.startswith("tmp"):
            try:
                xlsx_path.unlink()
            except:
                pass

    except Exception as e:
        results["error"] = str(e)
        results["details"]["exception"] = str(e)
        print(f"❌ Test failed with error: {e}")

    return results


def main():
    """Run the test."""
    results = run_test()

    # Save results
    output_file = Path(__file__).parent / "results.json"
    with open(output_file, "w") as f:
        json.dump(results, f, indent=2)

    print(f"Results saved to: {output_file}")
    return results


if __name__ == "__main__":
    main()
